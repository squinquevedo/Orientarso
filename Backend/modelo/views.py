import json
import csv
import re
import secrets
from pathlib import Path
from io import BytesIO, StringIO
from types import SimpleNamespace
from urllib.parse import urlencode

from django.conf import settings
from django.core.cache import cache
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.db import connection, transaction
from django.middleware.csrf import get_token
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from rest_framework import status
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response

from .models import EmailVerificationToken, PasswordResetToken, Perfil


ADMIN_EMAIL = 'admin@orientarso.com'
DOCUMENT_LENGTHS = {
    'CC': 10,
    'TI': 11,
    'CE': 7,
    'PAS': 9,
}
PENDING_REGISTRATION_TTL = 180
SIMULATED_ERROR_MESSAGES = {
    404: 'Error 404 simulado de forma controlada.',
    408: 'Error 408 simulado de forma controlada.',
    500: 'Error 500 simulado de forma controlada.',
    503: 'Error 503 simulado de forma controlada.',
}
RESPONSE_OPTIONS = [
    (1, 'Totalmente en desacuerdo', 1),
    (2, 'En desacuerdo', 2),
    (3, 'Ni de acuerdo, ni en desacuerdo', 3),
    (4, 'De acuerdo', 4),
    (5, 'Totalmente de acuerdo', 5),
]


def pending_registration_token_key(token):
    return f'pending_registration:token:{token}'


def pending_registration_email_key(email):
    return f'pending_registration:email:{email.lower()}'


def pending_registration_document_key(document):
    return f'pending_registration:document:{document}'


def verified_registration_token_key(token):
    return f'pending_registration:verified:{token}'


def clear_pending_registration(data):
    if not data:
        return
    cache.delete(pending_registration_token_key(data.get('token', '')))
    cache.delete(pending_registration_email_key(data.get('email', '')))
    cache.delete(pending_registration_document_key(data.get('numero_documento', '')))


def save_pending_registration(data):
    token = secrets.token_urlsafe(48)
    pending_data = {**data, 'token': token}
    previous_by_email = cache.get(pending_registration_email_key(data['email']))
    previous_by_document = cache.get(pending_registration_document_key(data['numero_documento']))
    clear_pending_registration(previous_by_email)
    clear_pending_registration(previous_by_document)
    cache.set(pending_registration_token_key(token), pending_data, PENDING_REGISTRATION_TTL)
    cache.set(pending_registration_email_key(data['email']), pending_data, PENDING_REGISTRATION_TTL)
    cache.set(pending_registration_document_key(data['numero_documento']), pending_data, PENDING_REGISTRATION_TTL)
    return pending_data


def build_verification_payload(user, profile=None):
    profile = profile or get_or_create_profile(user)
    payload = build_user_payload(user, profile)
    payload['message'] = 'Verificacion exitosa'
    return payload


def send_verification_email(user, token):
    verify_link = f'http://localhost:5173/verificar-email?token={token}'
    subject = 'Confirma tu correo - Orientarso'
    nombre = getattr(user, 'first_name', '') or getattr(user, 'nombre_completo', '') or 'Usuario'

    plain_message = f"""
Hola {nombre},

Gracias por registrarte en Orientarso.

Para activar tu cuenta, haz clic en el siguiente enlace:
{verify_link}

Este enlace expira en 3 minutos.

Si no creaste esta cuenta, ignora este mensaje.

Saludos,
El equipo de Orientarso
    """.strip()

    html_message = f"""
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin:0;padding:0;background-color:#f4f7fc;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f4f7fc;padding:40px 16px;">
    <tr>
      <td align="center">
        <table width="560" cellpadding="0" cellspacing="0" style="max-width:560px;width:100%;background-color:#ffffff;border-radius:16px;overflow:hidden;box-shadow:0 8px 30px rgba(0,0,0,0.08);">
          <tr>
            <td style="background:linear-gradient(135deg,#2563EB,#1d4ed8);padding:32px 40px;text-align:center;">
              <h1 style="margin:0;color:#ffffff;font-size:24px;font-weight:700;letter-spacing:0.5px;">
                Orientarso
              </h1>
              <p style="margin:6px 0 0;color:#bfdbfe;font-size:14px;">
                Orientación vocacional
              </p>
            </td>
          </tr>
          <tr>
            <td style="padding:40px 40px 32px;">
              <h2 style="margin:0 0 8px;color:#1F2937;font-size:20px;font-weight:600;">
                ¡Hola, {nombre}!
              </h2>
              <p style="margin:0 0 20px;color:#4b5563;font-size:15px;line-height:1.6;">
                Gracias por registrarte en <strong>Orientarso</strong>. 
                Para comenzar a descubrir tu vocación, solo falta un paso.
              </p>
              <p style="margin:0 0 24px;color:#4b5563;font-size:15px;line-height:1.6;">
                Confirma tu correo electrónico haciendo clic en el botón:
              </p>
              <table cellpadding="0" cellspacing="0" style="margin:0 auto 32px;">
                <tr>
                  <td align="center" style="background:linear-gradient(135deg,#2563EB,#1d4ed8);border-radius:8px;">
                    <a href="{verify_link}"
                       style="display:inline-block;padding:14px 40px;color:#ffffff;font-size:16px;font-weight:600;text-decoration:none;border-radius:8px;">
                      Verificar
                    </a>
                  </td>
                </tr>
              </table>
              <p style="margin:0 0 16px;color:#6b7280;font-size:13px;line-height:1.5;text-align:center;">
                O copia este enlace en tu navegador:
              </p>
              <p style="margin:0 0 24px;padding:12px 16px;background-color:#f3f4f6;border-radius:8px;font-size:12px;color:#374151;word-break:break-all;text-align:center;">
                {verify_link}
              </p>
              <hr style="border:none;border-top:1px solid #e5e7eb;margin:0 0 20px;">
              <p style="margin:0;color:#9ca3af;font-size:12px;line-height:1.5;text-align:center;">
                Este enlace expira en <strong>3 minutos</strong>.<br>
                Si no creaste esta cuenta, ignora este mensaje.
              </p>
            </td>
          </tr>
          <tr>
            <td style="background-color:#f9fafb;padding:20px 40px;text-align:center;border-top:1px solid #e5e7eb;">
              <p style="margin:0;color:#9ca3af;font-size:12px;">
                &copy; 2026 Orientarso &mdash; Todos los derechos reservados.
              </p>
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>
    """.strip()

    send_mail(
        subject=subject,
        message=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        html_message=html_message,
        fail_silently=False,
    )

    print(f"\n{'='*60}")
    print(f"  EMAIL DE VERIFICACION ENVIADO A: {user.email}")
    print(f"  Token: {token}")
    print(f"  Link:  {verify_link}")
    print(f"{'='*60}\n")


def send_password_reset_email(user, token):
    subject = 'Codigo de recuperacion - Orientarso'
    nombre = user.first_name or 'Usuario'

    plain_message = f"""
Hola {nombre},

Has solicitado restablecer tu contrasena en Orientarso.

Tu codigo de verificacion es: {token}

Este codigo expira en 3 minutos.

Si no solicitaste este cambio, ignora este mensaje.

Saludos,
El equipo de Orientarso
    """.strip()

    html_message = f"""
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin:0;padding:0;background-color:#f4f7fc;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f4f7fc;padding:40px 16px;">
    <tr>
      <td align="center">
        <table width="560" cellpadding="0" cellspacing="0" style="max-width:560px;width:100%;background-color:#ffffff;border-radius:16px;overflow:hidden;box-shadow:0 8px 30px rgba(0,0,0,0.08);">
          <tr>
            <td style="background:linear-gradient(135deg,#2563EB,#1d4ed8);padding:32px 40px;text-align:center;">
              <h1 style="margin:0;color:#ffffff;font-size:24px;font-weight:700;letter-spacing:0.5px;">
                Orientarso
              </h1>
              <p style="margin:6px 0 0;color:#bfdbfe;font-size:14px;">
                Recuperacion de contrasena
              </p>
            </td>
          </tr>
          <tr>
            <td style="padding:40px 40px 32px;">
              <h2 style="margin:0 0 8px;color:#1F2937;font-size:20px;font-weight:600;">
                ¡Hola, {nombre}!
              </h2>
              <p style="margin:0 0 20px;color:#4b5563;font-size:15px;line-height:1.6;">
                Recibimos una solicitud para restablecer tu contrasena en <strong>Orientarso</strong>.
              </p>
              <p style="margin:0 0 16px;color:#4b5563;font-size:15px;line-height:1.6;">
                Utiliza el siguiente codigo para restablecer tu contrasena:
              </p>
              <div style="text-align:center;margin:24px 0;padding:20px;background-color:#f3f4f6;border-radius:12px;letter-spacing:8px;font-size:32px;font-weight:700;color:#2563EB;">
                {token}
              </div>
              <p style="margin:0 0 8px;color:#6b7280;font-size:13px;line-height:1.5;text-align:center;">
                Este codigo expira en <strong>3 minutos</strong>.
              </p>
              <hr style="border:none;border-top:1px solid #e5e7eb;margin:24px 0 20px;">
              <p style="margin:0;color:#9ca3af;font-size:12px;line-height:1.5;text-align:center;">
                Si no solicitaste este cambio, ignora este mensaje.
              </p>
            </td>
          </tr>
          <tr>
            <td style="background-color:#f9fafb;padding:20px 40px;text-align:center;border-top:1px solid #e5e7eb;">
              <p style="margin:0;color:#9ca3af;font-size:12px;">
                &copy; 2026 Orientarso &mdash; Todos los derechos reservados.
              </p>
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>
    """.strip()

    send_mail(
        subject=subject,
        message=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        html_message=html_message,
        fail_silently=False,
    )

    print(f"\n{'='*60}")
    print(f"  EMAIL DE RECUPERACION ENVIADO A: {user.email}")
    print(f"  Codigo: {token}")
    print(f"{'='*60}\n")


def dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def ensure_response_options(cursor):
    cursor.execute('SELECT id, respuesta, valor FROM tb_respuesta ORDER BY valor, id')
    rows = cursor.fetchall()
    if list(rows) == RESPONSE_OPTIONS:
        return rows

    cursor.execute('DELETE FROM tb_respuesta')
    cursor.executemany(
        'INSERT INTO tb_respuesta (id, respuesta, valor) VALUES (%s, %s, %s)',
        RESPONSE_OPTIONS,
    )
    return RESPONSE_OPTIONS


def ensure_resultado_test_json_column(cursor):
    cursor.execute("SHOW COLUMNS FROM tb_resultado_test LIKE 'resultado_json'")
    if cursor.fetchone():
        return
    cursor.execute('ALTER TABLE tb_resultado_test ADD COLUMN resultado_json JSON NULL')


@api_view(['GET'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_simular_error(request, code):
    message = SIMULATED_ERROR_MESSAGES.get(code)
    if not message:
        return Response(
            {'error': 'Codigo de error no soportado para simulacion.'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    return Response({'error': message, 'code': code}, status=code)


def parse_json_request(request):
    if request.data:
        return request.data
    body = request.body.decode('utf-8') if request.body else '{}'
    return json.loads(body or '{}')


def normalize_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ['1', 'true', 'si', 'sí', 'yes', 'activo', 'activa']
    return bool(value)


def ensure_careers_active_column(cursor):
    cursor.execute("SHOW COLUMNS FROM carreras LIKE 'activa'")
    if cursor.fetchone():
        return
    cursor.execute('ALTER TABLE carreras ADD COLUMN activa TINYINT(1) NOT NULL DEFAULT 1')


def get_or_create_profile(user, tipo_documento='CC'):
    profile, _ = Perfil.objects.get_or_create(
        user=user,
        defaults={
            'tipo_documento': tipo_documento or 'CC',
            'rol': Perfil.ROL_ADMIN if (user.email or '').lower() == ADMIN_EMAIL else Perfil.ROL_ESTUDIANTE,
        },
    )

    expected_role = Perfil.ROL_ADMIN if (user.email or '').lower() == ADMIN_EMAIL else profile.rol
    updated_fields = []

    if not profile.tipo_documento:
        profile.tipo_documento = tipo_documento or 'CC'
        updated_fields.append('tipo_documento')

    if profile.rol != expected_role:
        profile.rol = expected_role
        updated_fields.append('rol')

    if updated_fields:
        profile.save(update_fields=updated_fields)

    return profile


def build_user_payload(user, profile=None):
    profile = profile or get_or_create_profile(user)
    is_admin = profile.rol == Perfil.ROL_ADMIN or (user.email or '').lower() == ADMIN_EMAIL
    foto_perfil_url = ''
    if profile.foto_perfil:
        foto_perfil_url = f'{settings.MEDIA_URL}{profile.foto_perfil}'
    return {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'tipo_documento': profile.tipo_documento,
        'foto_perfil': profile.foto_perfil,
        'foto_perfil_url': foto_perfil_url,
        'rol': Perfil.ROL_ADMIN if is_admin else profile.rol,
        'is_admin': is_admin,
        'is_authenticated': user.is_authenticated,
        'is_active': user.is_active,
    }


def ensure_admin_user(request):
    profile = get_or_create_profile(request.user)
    is_admin = profile.rol == Perfil.ROL_ADMIN or (request.user.email or '').lower() == ADMIN_EMAIL
    if not is_admin:
        return None, Response(
            {'error': 'No tienes permisos para acceder a este recurso'},
            status=status.HTTP_403_FORBIDDEN,
        )
    return profile, None


def serialize_university_rows():
    with connection.cursor() as cursor:
        cursor.execute(
            '''
            SELECT
                u.id,
                u.nombre,
                u.tipo,
                u.localidad,
                u.direccion,
                u.sitio_web
            FROM universidades u
            ORDER BY u.nombre
            '''
        )
        universities = dictfetchall(cursor)

        cursor.execute(
            '''
            SELECT
                cu.id,
                cu.id_universidad,
                cu.id_carrera,
                c.Nom_carrera AS nombre_carrera,
                c.ID_area AS id_area,
                cu.modalidad,
                cu.duracion_semestres,
                cu.valor_semestre,
                cu.jornada,
                cu.activa
            FROM carrera_universidad cu
            INNER JOIN carreras c ON c.ID = cu.id_carrera
            ORDER BY cu.id_universidad, c.Nom_carrera
            '''
        )
        carreras = dictfetchall(cursor)

        cursor.execute(
            '''
            SELECT
                b.id,
                b.id_universidad,
                b.id_carrera,
                c.Nom_carrera AS nombre_carrera,
                b.tipo_beneficio,
                b.descripcion,
                b.porcentaje_descuento
            FROM beneficios_carrera b
            INNER JOIN carreras c ON c.ID = b.id_carrera
            ORDER BY b.id_universidad, c.Nom_carrera, b.tipo_beneficio
            '''
        )
        beneficios = dictfetchall(cursor)

        cursor.execute(
            '''
            SELECT
                cv.id,
                cv.id_universidad,
                cv.id_carrera,
                c.Nom_carrera AS nombre_carrera,
                cv.id_entidad,
                e.nombre AS nombre_entidad,
                cv.nombre_convenio,
                cv.descripcion,
                cv.fecha_inicio,
                cv.fecha_fin,
                cv.vigente
            FROM convenios cv
            INNER JOIN carreras c ON c.ID = cv.id_carrera
            INNER JOIN entidades_apoyo e ON e.id = cv.id_entidad
            ORDER BY cv.id_universidad, cv.nombre_convenio
            '''
        )
        convenios = dictfetchall(cursor)

    by_university = {row['id']: row for row in universities}
    for university in universities:
        university['carreras'] = []
        university['beneficios_carrera'] = []
        university['convenios'] = []
        university['entidades_apoyo'] = []
        university['tiene_convenios'] = False
        university['tiene_beneficios_carrera'] = False
        university['tiene_entidades_apoyo'] = False

    for carrera in carreras:
        carrera['activa'] = bool(carrera['activa'])
        by_university[carrera['id_universidad']]['carreras'].append(carrera)

    for beneficio in beneficios:
        by_university[beneficio['id_universidad']]['beneficios_carrera'].append(beneficio)
        by_university[beneficio['id_universidad']]['tiene_beneficios_carrera'] = True

    entity_seen = {}
    for convenio in convenios:
        convenio['vigente'] = bool(convenio['vigente'])
        if convenio['fecha_inicio'] is not None:
            convenio['fecha_inicio'] = convenio['fecha_inicio'].isoformat()
        if convenio['fecha_fin'] is not None:
            convenio['fecha_fin'] = convenio['fecha_fin'].isoformat()

        university = by_university[convenio['id_universidad']]
        university['convenios'].append(convenio)
        university['tiene_convenios'] = True
        university['tiene_entidades_apoyo'] = True

        entity_seen.setdefault(university['id'], set())
        entity_key = convenio['id_entidad']
        if entity_key not in entity_seen[university['id']]:
            university['entidades_apoyo'].append({
                'id': convenio['id_entidad'],
                'nombre': convenio['nombre_entidad'],
            })
            entity_seen[university['id']].add(entity_key)

    return universities


def fetch_catalog_data():
    with connection.cursor() as cursor:
        ensure_careers_active_column(cursor)
        cursor.execute('SELECT id, nom_area FROM tb_area ORDER BY nom_area')
        areas = dictfetchall(cursor)

        cursor.execute(
            '''
            SELECT
                c.ID AS id,
                c.Nom_carrera AS nombre,
                c.ID_area AS id_area,
                a.nom_area AS area,
                c.activa
            FROM carreras c
            LEFT JOIN tb_area a ON a.id = c.ID_area
            ORDER BY c.Nom_carrera
            '''
        )
        carreras = dictfetchall(cursor)
        for carrera in carreras:
            carrera['activa'] = bool(carrera['activa'])

        cursor.execute(
            '''
            SELECT
                id,
                nombre,
                tipo,
                descripcion,
                contacto,
                sitio_web
            FROM entidades_apoyo
            ORDER BY nombre
            '''
        )
        entidades = dictfetchall(cursor)

    return {
        'areas': areas,
        'carreras_catalogo': carreras,
        'entidades_catalogo': entidades,
    }


def sync_university_relations(cursor, university_id, payload):
    carreras = payload.get('carreras') or []
    beneficios = payload.get('beneficios_carrera') or []
    convenios = payload.get('convenios') or []

    cursor.execute('DELETE FROM beneficios_carrera WHERE id_universidad = %s', [university_id])
    cursor.execute('DELETE FROM convenios WHERE id_universidad = %s', [university_id])
    cursor.execute('DELETE FROM carrera_universidad WHERE id_universidad = %s', [university_id])

    for carrera in carreras:
        id_carrera = carrera.get('id_carrera')
        if not id_carrera:
            continue
        cursor.execute(
            '''
            INSERT INTO carrera_universidad (
                id_carrera,
                id_universidad,
                modalidad,
                duracion_semestres,
                valor_semestre,
                jornada,
                activa
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''',
            [
                id_carrera,
                university_id,
                (carrera.get('modalidad') or '').strip() or None,
                carrera.get('duracion_semestres') or None,
                carrera.get('valor_semestre') or None,
                (carrera.get('jornada') or '').strip() or None,
                1 if normalize_bool(carrera.get('activa', True)) else 0,
            ],
        )

    for beneficio in beneficios:
        id_carrera = beneficio.get('id_carrera')
        tipo_beneficio = (beneficio.get('tipo_beneficio') or '').strip()
        if not id_carrera or not tipo_beneficio:
            continue
        cursor.execute(
            '''
            INSERT INTO beneficios_carrera (
                id_carrera,
                id_universidad,
                tipo_beneficio,
                descripcion,
                porcentaje_descuento
            ) VALUES (%s, %s, %s, %s, %s)
            ''',
            [
                id_carrera,
                university_id,
                tipo_beneficio,
                (beneficio.get('descripcion') or '').strip() or None,
                (beneficio.get('porcentaje_descuento') or '').strip() or None,
            ],
        )

    for convenio in convenios:
        id_carrera = convenio.get('id_carrera')
        id_entidad = convenio.get('id_entidad')
        nombre_convenio = (convenio.get('nombre_convenio') or '').strip()
        if not id_carrera or not id_entidad or not nombre_convenio:
            continue
        cursor.execute(
            '''
            INSERT INTO convenios (
                id_carrera,
                id_universidad,
                id_entidad,
                nombre_convenio,
                descripcion,
                fecha_inicio,
                fecha_fin,
                vigente
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ''',
            [
                id_carrera,
                university_id,
                id_entidad,
                nombre_convenio,
                (convenio.get('descripcion') or '').strip() or None,
                convenio.get('fecha_inicio') or None,
                convenio.get('fecha_fin') or None,
                1 if normalize_bool(convenio.get('vigente', True)) else 0,
            ],
        )


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_password_reset_request(request):
    try:
        data = json.loads(request.body)
        email = (data.get('email') or '').strip().lower()
        if not email:
            return Response({'error': 'Email requerido'}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.filter(email__iexact=email).first()
        if not user:
            return Response({'message': 'Si el correo existe, recibiras un codigo de recuperacion.'}, status=status.HTTP_200_OK)

        token = PasswordResetToken.generate_for_user(user)
        send_password_reset_email(user, token)

        return Response({'message': 'Codigo de recuperacion enviado a tu correo.'}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_password_reset_verify(request):
    try:
        data = json.loads(request.body)
        email = (data.get('email') or '').strip().lower()
        code = (data.get('code') or '').strip()
        new_password = (data.get('new_password') or '').strip()

        if not email or not code:
            return Response({'error': 'Email y codigo requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.filter(email__iexact=email).first()
        if not user:
            return Response({'error': 'Codigo invalido o expirado'}, status=status.HTTP_400_BAD_REQUEST)

        rt = PasswordResetToken.objects.filter(user=user, token=code, used=False).first()
        if not rt:
            return Response({'error': 'Codigo incorrecto'}, status=status.HTTP_400_BAD_REQUEST)

        if rt.is_expired():
            rt.used = True
            rt.save()
            return Response({'error': 'Codigo expirado'}, status=status.HTTP_410_GONE)

        if not new_password:
            return Response({'message': 'Codigo verificado correctamente.'}, status=status.HTTP_200_OK)

        if len(new_password) < 8 or len(new_password) > 16:
            return Response({'error': 'La contrasena debe tener entre 8 y 16 caracteres.'}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.save()
        rt.used = True
        rt.save()

        return Response({'message': 'Contrasena actualizada exitosamente.'}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_login(request):
    try:
        data = json.loads(request.body)
        identifier = (data.get('username') or '').strip()
        password = (data.get('password') or '').strip()

        user = None
        if '@' in identifier:
            user = User.objects.filter(email__iexact=identifier).first()
        if user is None:
            user = User.objects.filter(username__iexact=identifier).first()

        if user is None:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        if not user.is_active:
            return Response(
                {
                    'error': 'Cuenta no verificada. Revisa tu correo para activar tu cuenta.',
                    'requires_verification': True,
                    'email': user.email,
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        authenticated = authenticate(request, username=user.username, password=password)
        if authenticated is None:
            return Response({'error': 'Contrasena incorrecta'}, status=status.HTTP_401_UNAUTHORIZED)

        login(request, authenticated)
        profile = get_or_create_profile(authenticated)
        payload = build_user_payload(authenticated, profile)
        payload['message'] = 'Login exitoso'
        return Response(payload, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_registro(request):
    try:
        data = json.loads(request.body)
        tipo_documento = (data.get('tipo_documento') or '').strip()
        numero_documento = (data.get('numero_documento') or '').strip()
        nombre_completo = (data.get('nombre_completo') or '').strip()
        email = (data.get('email') or '').strip().lower()
        password1 = (data.get('password1') or '').strip()
        password2 = (data.get('password2') or '').strip()

        if not all([tipo_documento, numero_documento, nombre_completo, email, password1, password2]):
            return Response({'error': 'Todos los campos son requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        max_document_length = DOCUMENT_LENGTHS.get(tipo_documento)
        if not max_document_length:
            return Response({'error': 'Tipo de documento no valido'}, status=status.HTTP_400_BAD_REQUEST)

        if not numero_documento.isdigit():
            return Response({'error': 'El numero de documento solo puede contener numeros'}, status=status.HTTP_400_BAD_REQUEST)

        if len(numero_documento) > max_document_length:
            return Response(
                {'error': f'El numero de documento no puede superar {max_document_length} digitos'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if password1 != password2:
            return Response({'error': 'Las contraseñas no coinciden'}, status=status.HTTP_400_BAD_REQUEST)

        if len(password1) < 8 or len(password1) > 16:
            return Response({'error': 'La contrasena debe tener entre 8 y 16 caracteres.'}, status=status.HTTP_400_BAD_REQUEST)

        if not re.search(r'[A-Z]', password1):
            return Response({'error': 'La contrasena debe contener al menos una mayuscula.'}, status=status.HTTP_400_BAD_REQUEST)

        if not re.search(r'[^a-zA-Z0-9]', password1):
            return Response({'error': 'La contrasena debe contener al menos un caracter especial.'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(username__iexact=numero_documento).exists():
            return Response({'error': 'El numero de documento ya esta registrado'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email__iexact=email).exists():
            return Response({'error': 'El email ya está registrado'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            es_admin = email.lower() == ADMIN_EMAIL
            if es_admin:
                user = User.objects.create_user(
                    username=numero_documento,
                    email=email,
                    password=password1,
                    first_name=nombre_completo,
                    is_active=True,
                )
                profile = get_or_create_profile(user, tipo_documento=tipo_documento)
                profile.tipo_documento = tipo_documento
                profile.save(update_fields=['tipo_documento', 'rol'])
            else:
                pending = save_pending_registration({
                    'tipo_documento': tipo_documento,
                    'numero_documento': numero_documento,
                    'nombre_completo': nombre_completo,
                    'email': email,
                    'password': make_password(password1),
                })
                try:
                    send_verification_email(SimpleNamespace(**pending), pending['token'])
                except Exception:
                    clear_pending_registration(pending)
                    raise
                user = None
                profile = None

        payload = {
            'message': 'Te enviamos un correo de verificacion. Tu cuenta se creara cuando confirmes el enlace.'
                       if not es_admin else
                       'Usuario administrador creado exitosamente.',
            'username': user.username if user else numero_documento,
            'email': email,
            'rol': profile.rol if profile else Perfil.ROL_ESTUDIANTE,
            'requires_verification': not es_admin,
        }
        return Response(payload, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_logout(request):
    logout(request)
    return Response({'message': 'Logout exitoso'}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_user(request):
    payload = build_user_payload(request.user)
    if payload.get('foto_perfil_url'):
        payload['foto_perfil_url'] = request.build_absolute_uri(payload['foto_perfil_url'])
    return Response(payload, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_user_profile_photo(request):
    profile = get_or_create_profile(request.user)

    if request.method == 'DELETE':
        if profile.foto_perfil:
            old_path = settings.MEDIA_ROOT / profile.foto_perfil
            if old_path.exists():
                old_path.unlink()
        profile.foto_perfil = ''
        profile.save(update_fields=['foto_perfil'])
        return Response({'foto_perfil_url': ''}, status=status.HTTP_200_OK)

    uploaded_file = request.FILES.get('foto') or request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'No se envio ninguna foto'}, status=status.HTTP_400_BAD_REQUEST)

    if uploaded_file.size > 1024 * 1024:
        return Response({'error': 'La foto debe pesar maximo 1 MB'}, status=status.HTTP_400_BAD_REQUEST)

    content_type = (uploaded_file.content_type or '').lower()
    if content_type and not content_type.startswith('image/'):
        return Response({'error': 'El archivo debe ser una imagen'}, status=status.HTTP_400_BAD_REQUEST)

    photo_dir = settings.MEDIA_ROOT / 'user_photos'
    photo_dir.mkdir(parents=True, exist_ok=True)

    relative_path = Path('user_photos') / f'user_{request.user.id}.webp'
    absolute_path = settings.MEDIA_ROOT / relative_path

    if profile.foto_perfil and profile.foto_perfil != relative_path.as_posix():
        old_path = settings.MEDIA_ROOT / profile.foto_perfil
        if old_path.exists():
            old_path.unlink()

    with absolute_path.open('wb') as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)

    profile.foto_perfil = relative_path.as_posix()
    profile.save(update_fields=['foto_perfil'])

    return Response(
        {
            'foto_perfil': profile.foto_perfil,
            'foto_perfil_url': request.build_absolute_uri(f'{settings.MEDIA_URL}{profile.foto_perfil}'),
        },
        status=status.HTTP_200_OK,
    )


@ensure_csrf_cookie
@api_view(['GET'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_csrf(request):
    return Response({'csrfToken': get_token(request)}, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_verificar_email(request):
    if request.method == 'GET':
        token = request.query_params.get('token', '').strip()
        if not token:
            return Response({'error': 'No se pudo verificar la cuenta, intenta de nuevo.'}, status=status.HTTP_400_BAD_REQUEST)

        pending = cache.get(pending_registration_token_key(token))
        verified_user_id = cache.get(verified_registration_token_key(token))
        if verified_user_id:
            user = User.objects.filter(id=verified_user_id, is_active=True).first()
            if user:
                return Response(build_verification_payload(user), status=status.HTTP_200_OK)

        if pending:
            existing_user = User.objects.filter(
                username__iexact=pending['numero_documento'],
                email__iexact=pending['email'],
                is_active=True,
            ).first()
            if existing_user:
                clear_pending_registration(pending)
                cache.set(verified_registration_token_key(token), existing_user.id, 600)
                return Response(build_verification_payload(existing_user), status=status.HTTP_200_OK)

            if User.objects.filter(username__iexact=pending['numero_documento']).exists():
                clear_pending_registration(pending)
                return Response({'error': 'El numero de documento ya esta registrado'}, status=status.HTTP_400_BAD_REQUEST)

            if User.objects.filter(email__iexact=pending['email']).exists():
                clear_pending_registration(pending)
                return Response({'error': 'El email ya esta registrado'}, status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                user = User(
                    username=pending['numero_documento'],
                    email=pending['email'],
                    first_name=pending['nombre_completo'],
                    is_active=True,
                )
                user.password = pending['password']
                user.save()
                profile = get_or_create_profile(user, tipo_documento=pending['tipo_documento'])
                profile.tipo_documento = pending['tipo_documento']
                profile.rol = Perfil.ROL_ESTUDIANTE
                profile.save(update_fields=['tipo_documento', 'rol'])
                clear_pending_registration(pending)
                cache.set(verified_registration_token_key(token), user.id, 600)

            payload = build_user_payload(user, profile)
            payload['message'] = 'VerificaciÃ³n exitosa'
            return Response(payload, status=status.HTTP_200_OK)

        try:
            vt = EmailVerificationToken.objects.select_related('user').get(token=token)
        except EmailVerificationToken.DoesNotExist:
            return Response({'error': 'No se pudo verificar la cuenta, intenta de nuevo.'}, status=status.HTTP_404_NOT_FOUND)

        if vt.is_expired():
            vt.delete()
            return Response({'error': 'No se pudo verificar la cuenta, intenta de nuevo.'}, status=status.HTTP_410_GONE)

        user = vt.user
        if user.is_active:
            vt.delete()
        else:
            user.is_active = True
            user.save(update_fields=['is_active'])
            vt.delete()

        profile = get_or_create_profile(user)
        payload = build_user_payload(user, profile)
        payload['message'] = 'Verificación exitosa'

        return Response(payload, status=status.HTTP_200_OK)

    try:
        data = json.loads(request.body)
        email = (data.get('email') or '').strip().lower()
        if not email:
            return Response({'error': 'Email requerido'}, status=status.HTTP_400_BAD_REQUEST)

        pending = cache.get(pending_registration_email_key(email))
        if pending:
            pending = save_pending_registration({
                'tipo_documento': pending['tipo_documento'],
                'numero_documento': pending['numero_documento'],
                'nombre_completo': pending['nombre_completo'],
                'email': pending['email'],
                'password': pending['password'],
            })
            try:
                send_verification_email(SimpleNamespace(**pending), pending['token'])
            except Exception:
                clear_pending_registration(pending)
                raise
            return Response({'message': 'Correo de verificacion reenviado. Revisa tu bandeja de entrada.'}, status=status.HTTP_200_OK)

        user = User.objects.filter(email__iexact=email, is_active=False).first()
        if not user:
            return Response({'error': 'No hay una cuenta pendiente de verificacion con ese email'}, status=status.HTTP_404_NOT_FOUND)

        token = EmailVerificationToken.generate_for_user(user)
        send_verification_email(user, token)

        return Response({'message': 'Correo de verificacion reenviado. Revisa tu bandeja de entrada.'}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_preguntas(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT id, pregunta, valor, id_area FROM tb_preguntas ORDER BY id')
            pregunta_rows = cursor.fetchall()
            respuesta_rows = ensure_response_options(cursor)

        preguntas = [
            {
                'id': row[0],
                'pregunta': row[1],
                'valor': row[2],
                'id_area': row[3],
            }
            for row in pregunta_rows
        ]
        opciones = [
            {
                'id': row[0],
                'label': row[1],
                'value': row[2],
            }
            for row in respuesta_rows
        ]
        for pregunta in preguntas:
            pregunta['opciones'] = opciones
        return Response(preguntas, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
@authentication_classes([])
def api_areas(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT id, nom_area FROM tb_area ORDER BY id')
            rows = cursor.fetchall()

        areas = [{'id': row[0], 'nom_area': row[1]} for row in rows]
        return Response(areas, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_university_report_data(request):
    try:
        return Response(
            {
                'universities': serialize_university_rows(),
                **fetch_catalog_data(),
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def build_test_attempts_payload(user_id):
    with connection.cursor() as cursor:
        cursor.execute(
            '''
            SELECT id, fecha_registro, intento
            FROM tb_test
            WHERE id_auth_user = %s
            ORDER BY intento DESC, fecha_registro DESC, id DESC
            ''',
            [user_id],
        )
        tests = dictfetchall(cursor)

        if not tests:
            return []

        test_ids = [test['id'] for test in tests]
        placeholders = ', '.join(['%s'] * len(test_ids))
        cursor.execute(
            f'''
            SELECT
                rt.id_test,
                rt.id_area,
                a.nom_area,
                rt.puntaje_total AS score,
                COALESCE(maximos.max_score, 0) AS max_score
            FROM tb_resultado_test rt
            LEFT JOIN tb_area a ON a.id = rt.id_area
            LEFT JOIN (
                SELECT id_area, SUM(valor) AS max_score
                FROM tb_preguntas
                GROUP BY id_area
            ) maximos ON maximos.id_area = rt.id_area
            WHERE rt.id_test IN ({placeholders})
            ORDER BY rt.id_test, score DESC
            ''',
            test_ids,
        )
        rows = dictfetchall(cursor)

    results_by_test = {test['id']: [] for test in tests}
    for row in rows:
        score = float(row['score'] or 0)
        max_score = float(row['max_score'] or 0)
        pct = round((score / max_score) * 100) if max_score else 0
        results_by_test[row['id_test']].append(
            {
                'areaKey': str(row['id_area']),
                'areaName': row['nom_area'] or f"Area {row['id_area']}",
                'score': score,
                'max': max_score,
                'pct': pct,
            }
        )

    attempts = []
    for test in tests:
        fecha = test['fecha_registro']
        attempts.append(
            {
                'id_test': test['id'],
                'intento': test['intento'],
                'fecha_registro': fecha.isoformat() if fecha else '',
                'resultados': sorted(
                    results_by_test.get(test['id'], []),
                    key=lambda item: item['pct'],
                    reverse=True,
                ),
            }
        )

    return attempts


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_test(request):
    if request.method == 'GET':
        try:
            return Response({'attempts': build_test_attempts_payload(request.user.id)}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    try:
        data = request.data or {}
        respuestas = data.get('respuestas') or []

        if not isinstance(respuestas, list):
            return Response({'error': 'Formato de respuestas invalido'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            with connection.cursor() as cursor:
                ensure_resultado_test_json_column(cursor)
                cursor.execute('SELECT COUNT(*) FROM tb_test WHERE id_auth_user = %s', [request.user.id])
                count = cursor.fetchone()[0] or 0
                intento = count + 1

                cursor.execute('SELECT MAX(id) FROM tb_test')
                max_id = cursor.fetchone()[0] or 0
                test_id = max_id + 1
                fecha_registro = timezone.now()

                cursor.execute(
                    'INSERT INTO tb_test (id, id_auth_user, fecha_registro, intento) VALUES (%s, %s, %s, %s)',
                    [test_id, request.user.id, fecha_registro, intento],
                )

                valores = []
                for item in respuestas:
                    try:
                        id_pregunta = int(item.get('id_pregunta'))
                        respuesta = float(item.get('respuesta'))
                    except (TypeError, ValueError):
                        continue
                    valores.append((test_id, id_pregunta, respuesta))

                if valores:
                    opciones_catalogo = ensure_response_options(cursor)
                    opciones_validas = {float(row[2]) for row in opciones_catalogo}
                    max_opcion = max(opciones_validas) if opciones_validas else 5

                    pregunta_ids = [id_pregunta for _, id_pregunta, _ in valores]
                    placeholders = ', '.join(['%s'] * len(pregunta_ids))
                    cursor.execute(
                        f'''
                        SELECT p.id, p.pregunta, p.id_area, a.nom_area, p.valor
                        FROM tb_preguntas p
                        LEFT JOIN tb_area a ON a.id = p.id_area
                        WHERE p.id IN ({placeholders})
                        ''',
                        pregunta_ids,
                    )
                    preguntas_por_id = {
                        int(row[0]): {
                            'pregunta': row[1],
                            'id_area': int(row[2]),
                            'area': row[3] or f"Area {row[2]}",
                            'valor': float(row[4] or 0),
                        }
                        for row in cursor.fetchall()
                    }
                    opciones_por_valor = {
                        float(row[2]): {
                            'id': int(row[0]),
                            'respuesta': row[1],
                            'valor': int(row[2]),
                        }
                        for row in opciones_catalogo
                    }

                    puntajes_por_area = {}
                    respuestas_detalle = []
                    for _, id_pregunta_val, respuesta_val in valores:
                        if respuesta_val not in opciones_validas:
                            continue
                        pregunta = preguntas_por_id.get(id_pregunta_val)
                        if not pregunta:
                            continue
                        id_area = pregunta['id_area']
                        puntaje = (respuesta_val / max_opcion) * pregunta['valor']
                        puntajes_por_area[id_area] = puntajes_por_area.get(id_area, 0) + puntaje
                        opcion = opciones_por_valor.get(respuesta_val, {})
                        respuestas_detalle.append(
                            {
                                'id_pregunta': id_pregunta_val,
                                'pregunta': pregunta['pregunta'],
                                'id_area': id_area,
                                'area': pregunta['area'],
                                'respuesta': opcion.get('respuesta', ''),
                                'valor_respuesta': int(respuesta_val),
                                'puntaje_pregunta': round(puntaje, 2),
                            }
                        )

                    puntajes_detalle = [
                        {
                            'id_area': id_area,
                            'area': next(
                                (
                                    pregunta['area']
                                    for pregunta in preguntas_por_id.values()
                                    if pregunta['id_area'] == id_area
                                ),
                                f'Area {id_area}',
                            ),
                            'puntaje_total': round(puntaje_total, 2),
                        }
                        for id_area, puntaje_total in puntajes_por_area.items()
                    ]
                    resultado_json = json.dumps(
                        {
                            'id_test': test_id,
                            'usuario': {
                                'id': request.user.id,
                                'username': request.user.username,
                                'email': request.user.email,
                                'nombre': request.user.get_full_name() or request.user.first_name or request.user.username,
                            },
                            'fecha_registro': fecha_registro.isoformat(),
                            'intento': intento,
                            'respuestas': respuestas_detalle,
                            'puntajes_por_area': puntajes_detalle,
                        },
                        ensure_ascii=False,
                    )

                    cursor.execute('SELECT MAX(id) FROM tb_resultado_test')
                    max_resultado_id = cursor.fetchone()[0] or 0
                    resultados_con_id = []
                    for idx, (id_area, puntaje_total) in enumerate(puntajes_por_area.items(), start=1):
                        resultados_con_id.append((max_resultado_id + idx, test_id, id_area, puntaje_total, resultado_json))

                    if resultados_con_id:
                        cursor.executemany(
                            '''
                            INSERT INTO tb_resultado_test
                                (id, id_test, id_area, puntaje_total, resultado_json)
                            VALUES (%s, %s, %s, %s, %s)
                            ''',
                            resultados_con_id,
                        )

        return Response(
            {
                'id_test': test_id,
                'intento': intento,
                'attempts': build_test_attempts_payload(request.user.id),
            },
            status=status.HTTP_201_CREATED,
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_admin_summary(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    return Response(
        {
            'users': [build_user_payload(user) for user in User.objects.order_by('id')],
            'universities': serialize_university_rows(),
            **fetch_catalog_data(),
        },
        status=status.HTTP_200_OK,
    )


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_admin_user_detail(request, user_id):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        if user.id == request.user.id:
            return Response({'error': 'No puedes eliminar tu propio usuario admin'}, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            with connection.cursor() as cursor:
                cursor.execute('DELETE FROM django_admin_log WHERE user_id = %s', [user.id])
                cursor.execute('DELETE FROM auth_user_groups WHERE user_id = %s', [user.id])
                cursor.execute('DELETE FROM auth_user_user_permissions WHERE user_id = %s', [user.id])
                cursor.execute('DELETE FROM usuarios_perfil WHERE user_id = %s', [user.id])
                cursor.execute(
                    '''
                    DELETE FROM tb_resultado_test
                    WHERE id_test IN (
                        SELECT id FROM tb_test WHERE id_auth_user = %s
                    )
                    ''',
                    [user.id],
                )
                cursor.execute('DELETE FROM tb_test WHERE id_auth_user = %s', [user.id])
            user.delete()
        return Response({'message': 'Usuario eliminado'}, status=status.HTTP_200_OK)

    data = parse_json_request(request)
    email = (data.get('email') or user.email).strip().lower()
    if not email:
        return Response({'error': 'El email es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.exclude(pk=user.id).filter(email__iexact=email).exists():
        return Response({'error': 'Ya existe otro usuario con ese email'}, status=status.HTTP_400_BAD_REQUEST)

    username = (data.get('username') or user.username).strip()
    if not username:
        return Response({'error': 'El numero de documento es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.exclude(pk=user.id).filter(username__iexact=username).exists():
        return Response({'error': 'Ya existe otro usuario con ese numero de documento'}, status=status.HTTP_400_BAD_REQUEST)

    rol = (data.get('rol') or Perfil.ROL_ESTUDIANTE).strip().lower()
    if email == ADMIN_EMAIL:
        rol = Perfil.ROL_ADMIN
    elif rol not in [Perfil.ROL_ADMIN, Perfil.ROL_ESTUDIANTE]:
        rol = Perfil.ROL_ESTUDIANTE

    user.email = email
    user.username = username
    user.first_name = (data.get('first_name') or user.first_name).strip()
    if 'is_active' in data:
        user.is_active = normalize_bool(data.get('is_active'))
    user.save()

    profile = get_or_create_profile(user, tipo_documento=(data.get('tipo_documento') or 'CC').strip())
    profile.tipo_documento = (data.get('tipo_documento') or profile.tipo_documento or 'CC').strip()
    profile.rol = rol
    profile.save(update_fields=['tipo_documento', 'rol'])

    return Response(build_user_payload(user, profile), status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_career_create(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    data = parse_json_request(request)
    nombre = (data.get('nombre') or '').strip()
    id_area = data.get('id_area')
    activa = normalize_bool(data.get('activa', True))
    if not nombre or not id_area:
        return Response({'error': 'Nombre e area son obligatorios'}, status=status.HTTP_400_BAD_REQUEST)

    with connection.cursor() as cursor:
        ensure_careers_active_column(cursor)
        cursor.execute(
            'INSERT INTO carreras (Nom_carrera, ID_area, activa) VALUES (%s, %s, %s)',
            [nombre, id_area, 1 if activa else 0],
        )
        career_id = cursor.lastrowid

    return Response({'id': career_id, 'nombre': nombre, 'id_area': id_area, 'activa': activa}, status=status.HTTP_201_CREATED)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_admin_career_detail(request, career_id):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    with connection.cursor() as cursor:
        ensure_careers_active_column(cursor)
        if request.method == 'DELETE':
            cursor.execute('UPDATE carreras SET activa = 0 WHERE ID = %s', [career_id])
            cursor.execute('UPDATE carrera_universidad SET activa = 0 WHERE id_carrera = %s', [career_id])
            return Response({'message': 'Carrera desactivada'}, status=status.HTTP_200_OK)

        data = parse_json_request(request)
        nombre = (data.get('nombre') or '').strip()
        id_area = data.get('id_area')
        activa = normalize_bool(data.get('activa', True))
        if not nombre or not id_area:
            return Response({'error': 'Nombre e area son obligatorios'}, status=status.HTTP_400_BAD_REQUEST)

        cursor.execute(
            'UPDATE carreras SET Nom_carrera = %s, ID_area = %s, activa = %s WHERE ID = %s',
            [nombre, id_area, 1 if activa else 0, career_id],
        )

    return Response({'id': career_id, 'nombre': nombre, 'id_area': id_area, 'activa': activa}, status=status.HTTP_200_OK)


def parse_career_bulk_file(uploaded_file):
    filename = (uploaded_file.name or '').lower()
    if filename.endswith('.csv'):
        content = uploaded_file.read().decode('utf-8-sig')
        return list(csv.DictReader(StringIO(content)))

    if filename.endswith('.xlsx'):
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or '').strip().lower() for cell in rows[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
        ]

    raise ValueError('El archivo debe ser CSV o XLSX')


def get_normalized_value(normalized_row, *keys):
    for key in keys:
        value = normalized_row.get(key)
        if value not in [None, '']:
            return value
    return ''


def build_catalog_lookup(cursor):
    cursor.execute('SELECT ID AS id, Nom_carrera AS nombre FROM carreras')
    careers = dictfetchall(cursor)
    cursor.execute('SELECT id, nombre FROM entidades_apoyo')
    entities = dictfetchall(cursor)
    return {
        'career_by_id': {str(item['id']): item for item in careers},
        'career_by_name': {str(item['nombre']).strip().lower(): item for item in careers},
        'entity_by_id': {str(item['id']): item for item in entities},
        'entity_by_name': {str(item['nombre']).strip().lower(): item for item in entities},
    }


def find_catalog_item(normalized, by_id, by_name, id_keys, name_keys):
    item_id = get_normalized_value(normalized, *id_keys)
    item_name = str(get_normalized_value(normalized, *name_keys)).strip()
    if item_id not in [None, '']:
        item = by_id.get(str(item_id).strip())
        if item:
            return item
    if item_name:
        return by_name.get(item_name.lower())
    return None


def normalize_date_cell(value):
    if not value:
        return ''
    if hasattr(value, 'date') and callable(value.date):
        return value.date().isoformat()
    if hasattr(value, 'isoformat') and callable(value.isoformat):
        return value.isoformat()
    return str(value).strip()


def parse_career_catalog_rows(rows):
    parsed_rows = []
    rejected_rows = []

    with connection.cursor() as cursor:
        ensure_careers_active_column(cursor)
        cursor.execute('SELECT id, nom_area FROM tb_area')
        areas = dictfetchall(cursor)
        cursor.execute('SELECT ID AS id, LOWER(Nom_carrera) AS nombre FROM carreras')
        existing_careers = dictfetchall(cursor)

    area_by_id = {str(row['id']): row for row in areas}
    area_by_name = {str(row['nom_area']).strip().lower(): row for row in areas}
    existing_by_name = {row['nombre']: row['id'] for row in existing_careers}

    def append_rejected(normalized_row, observation):
        rejected_rows.append(
            {
                'nombre': str(get_normalized_value(normalized_row, 'nombre', 'carrera')).strip(),
                'area': str(get_normalized_value(normalized_row, 'area', 'área', 'id_area', 'area_id')).strip(),
                'estado': str(get_normalized_value(normalized_row, 'estado', 'activa')).strip() or 'activa',
                'observacion': observation,
            }
        )

    for row in rows:
        normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
        nombre = str(get_normalized_value(normalized, 'nombre', 'carrera')).strip()
        id_area = get_normalized_value(normalized, 'id_area', 'area_id')
        area_name = str(get_normalized_value(normalized, 'area', 'área')).strip().lower()
        area = None

        if id_area not in [None, '']:
            area = area_by_id.get(str(id_area).strip())
        if not area and area_name:
            area = area_by_name.get(area_name)

        missing_fields = []
        if not nombre:
            missing_fields.append('nombre')
        if not area:
            missing_fields.append('area')
        if missing_fields:
            append_rejected(normalized, f"Campos obligatorios faltantes o invalidos: {', '.join(missing_fields)}.")
            continue

        action = 'actualizar' if nombre.lower() in existing_by_name else 'crear'
        parsed_rows.append(
            {
                'nombre': nombre,
                'id_area': area['id'],
                'area': area['nom_area'],
                'activa': normalize_bool(get_normalized_value(normalized, 'estado', 'activa') or True),
                'accion': action,
            }
        )

    return parsed_rows, rejected_rows


def save_career_catalog_rows(rows):
    created = 0
    updated = 0

    with connection.cursor() as cursor:
        ensure_careers_active_column(cursor)
        for row in rows:
            nombre = (row.get('nombre') or row.get('carrera') or '').strip()
            id_area = row.get('id_area') or row.get('area_id')
            activa = normalize_bool(row.get('activa', row.get('estado', True)))
            if not nombre or not id_area:
                continue

            cursor.execute('SELECT ID FROM carreras WHERE LOWER(Nom_carrera) = LOWER(%s)', [nombre])
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    'UPDATE carreras SET ID_area = %s, activa = %s WHERE ID = %s',
                    [id_area, 1 if activa else 0, existing[0]],
                )
                updated += 1
            else:
                cursor.execute(
                    'INSERT INTO carreras (Nom_carrera, ID_area, activa) VALUES (%s, %s, %s)',
                    [nombre, id_area, 1 if activa else 0],
                )
                created += 1

    return created, updated


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_university_career_link_preview(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    uploaded_file = request.FILES.get('archivo') or request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'Debes seleccionar un archivo CSV o XLSX'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = parse_career_bulk_file(uploaded_file)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    with connection.cursor() as cursor:
        cursor.execute('SELECT ID AS id, Nom_carrera AS nombre FROM carreras')
        careers = dictfetchall(cursor)

    career_by_id = {str(item['id']): item for item in careers}
    career_by_name = {str(item['nombre']).strip().lower(): item for item in careers}
    parsed_rows = []
    rejected_rows = []

    def append_rejected(normalized_row, observation):
        rejected_rows.append(
            {
                'carrera': str(get_normalized_value(normalized_row, 'carrera', 'nombre', 'id_carrera', 'id carrera', 'career_id')).strip(),
                'modalidad': str(get_normalized_value(normalized_row, 'modalidad')).strip(),
                'duracion_semestres': str(get_normalized_value(normalized_row, 'duracion_semestres', 'duracion semestres', 'duracion')).strip(),
                'valor_semestre': str(get_normalized_value(normalized_row, 'valor_semestre', 'valor semestre', 'valor')).strip(),
                'jornada': str(get_normalized_value(normalized_row, 'jornada')).strip(),
                'estado': str(get_normalized_value(normalized_row, 'estado', 'activa')).strip() or 'activa',
                'observacion': observation,
            }
        )

    for row in rows:
        normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
        id_carrera = normalized.get('id_carrera') or normalized.get('id carrera') or normalized.get('career_id')
        career_name = str(normalized.get('carrera') or normalized.get('nombre') or '').strip()
        career = None

        if id_carrera not in [None, '']:
            career = career_by_id.get(str(id_carrera).strip())
        if not career and career_name:
            career = career_by_name.get(career_name.lower())

        if not career:
            append_rejected(
                normalized,
                'Carrera no encontrada. Por favor vincule la carrera en la BD.',
            )
            continue

        modalidad = str(get_normalized_value(normalized, 'modalidad')).strip()
        duracion_semestres = str(get_normalized_value(normalized, 'duracion_semestres', 'duracion semestres', 'duracion')).strip()
        valor_semestre = str(get_normalized_value(normalized, 'valor_semestre', 'valor semestre', 'valor')).strip()
        jornada = str(get_normalized_value(normalized, 'jornada')).strip()
        missing_fields = []
        if not modalidad:
            missing_fields.append('modalidad')
        if not duracion_semestres:
            missing_fields.append('duracion_semestres')
        if not valor_semestre:
            missing_fields.append('valor_semestre')
        if not jornada:
            missing_fields.append('jornada')
        if missing_fields:
            append_rejected(normalized, f"Campos obligatorios faltantes: {', '.join(missing_fields)}.")
            continue

        parsed_rows.append(
            {
                'id_carrera': career['id'],
                'carrera': career['nombre'],
                'modalidad': modalidad,
                'duracion_semestres': duracion_semestres,
                'valor_semestre': valor_semestre,
                'jornada': jornada,
                'activa': normalize_bool(normalized.get('estado', normalized.get('activa', True))),
            }
        )

    return Response(
        {
            'rows': parsed_rows,
            'rejected_rows': rejected_rows,
            'loaded': len(parsed_rows),
            'rejected': len(rejected_rows),
            'skipped': len(rejected_rows),
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_university_benefit_preview(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    uploaded_file = request.FILES.get('archivo') or request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'Debes seleccionar un archivo CSV o XLSX'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = parse_career_bulk_file(uploaded_file)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    with connection.cursor() as cursor:
        catalogs = build_catalog_lookup(cursor)

    parsed_rows = []
    rejected_rows = []

    def append_rejected(normalized_row, observation):
        rejected_rows.append(
            {
                'carrera': str(get_normalized_value(normalized_row, 'carrera', 'nombre_carrera', 'id_carrera', 'id carrera')).strip(),
                'tipo_beneficio': str(get_normalized_value(normalized_row, 'tipo_beneficio', 'tipo beneficio', 'beneficio')).strip(),
                'porcentaje_descuento': str(get_normalized_value(normalized_row, 'porcentaje_descuento', 'porcentaje descuento', 'porcentaje')).strip(),
                'descripcion': str(get_normalized_value(normalized_row, 'descripcion', 'descripción')).strip(),
                'observacion': observation,
            }
        )

    for row in rows:
        normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
        career = find_catalog_item(
            normalized,
            catalogs['career_by_id'],
            catalogs['career_by_name'],
            ['id_carrera', 'id carrera', 'career_id'],
            ['carrera', 'nombre_carrera', 'nombre'],
        )
        tipo_beneficio = str(get_normalized_value(normalized, 'tipo_beneficio', 'tipo beneficio', 'beneficio')).strip()

        if not career:
            append_rejected(normalized, 'Carrera no encontrada. Por favor vincule la carrera en la BD.')
            continue
        if not tipo_beneficio:
            append_rejected(normalized, 'Campo obligatorio faltante: tipo_beneficio.')
            continue

        parsed_rows.append(
            {
                'id_carrera': career['id'],
                'carrera': career['nombre'],
                'tipo_beneficio': tipo_beneficio,
                'porcentaje_descuento': str(get_normalized_value(normalized, 'porcentaje_descuento', 'porcentaje descuento', 'porcentaje')).strip(),
                'descripcion': str(get_normalized_value(normalized, 'descripcion', 'descripción')).strip(),
            }
        )

    return Response(
        {
            'rows': parsed_rows,
            'rejected_rows': rejected_rows,
            'loaded': len(parsed_rows),
            'rejected': len(rejected_rows),
            'skipped': len(rejected_rows),
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_university_agreement_preview(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    uploaded_file = request.FILES.get('archivo') or request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'Debes seleccionar un archivo CSV o XLSX'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = parse_career_bulk_file(uploaded_file)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    with connection.cursor() as cursor:
        catalogs = build_catalog_lookup(cursor)

    parsed_rows = []
    rejected_rows = []

    def append_rejected(normalized_row, observation):
        rejected_rows.append(
            {
                'carrera': str(get_normalized_value(normalized_row, 'carrera', 'nombre_carrera', 'id_carrera', 'id carrera')).strip(),
                'entidad': str(get_normalized_value(normalized_row, 'entidad', 'entidad_apoyo', 'id_entidad', 'id entidad')).strip(),
                'nombre_convenio': str(get_normalized_value(normalized_row, 'nombre_convenio', 'nombre convenio', 'convenio')).strip(),
                'fecha_inicio': normalize_date_cell(get_normalized_value(normalized_row, 'fecha_inicio', 'fecha inicio')),
                'fecha_fin': normalize_date_cell(get_normalized_value(normalized_row, 'fecha_fin', 'fecha fin')),
                'vigente': str(get_normalized_value(normalized_row, 'vigente', 'estado')).strip() or 'vigente',
                'descripcion': str(get_normalized_value(normalized_row, 'descripcion', 'descripción')).strip(),
                'observacion': observation,
            }
        )

    for row in rows:
        normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
        career = find_catalog_item(
            normalized,
            catalogs['career_by_id'],
            catalogs['career_by_name'],
            ['id_carrera', 'id carrera', 'career_id'],
            ['carrera', 'nombre_carrera', 'nombre_carrera'],
        )
        entity = find_catalog_item(
            normalized,
            catalogs['entity_by_id'],
            catalogs['entity_by_name'],
            ['id_entidad', 'id entidad', 'entity_id'],
            ['entidad', 'entidad_apoyo', 'nombre_entidad'],
        )
        nombre_convenio = str(get_normalized_value(normalized, 'nombre_convenio', 'nombre convenio', 'convenio')).strip()

        missing_fields = []
        if not career:
            missing_fields.append('carrera no encontrada')
        if not entity:
            missing_fields.append('entidad no encontrada')
        if not nombre_convenio:
            missing_fields.append('nombre_convenio')
        if missing_fields:
            append_rejected(normalized, f"Campos invalidos o faltantes: {', '.join(missing_fields)}.")
            continue

        parsed_rows.append(
            {
                'id_carrera': career['id'],
                'carrera': career['nombre'],
                'id_entidad': entity['id'],
                'entidad': entity['nombre'],
                'nombre_convenio': nombre_convenio,
                'fecha_inicio': normalize_date_cell(get_normalized_value(normalized, 'fecha_inicio', 'fecha inicio')),
                'fecha_fin': normalize_date_cell(get_normalized_value(normalized, 'fecha_fin', 'fecha fin')),
                'vigente': normalize_bool(get_normalized_value(normalized, 'vigente', 'estado') or True),
                'descripcion': str(get_normalized_value(normalized, 'descripcion', 'descripción')).strip(),
            }
        )

    return Response(
        {
            'rows': parsed_rows,
            'rejected_rows': rejected_rows,
            'loaded': len(parsed_rows),
            'rejected': len(rejected_rows),
            'skipped': len(rejected_rows),
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_career_bulk_preview(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    uploaded_file = request.FILES.get('archivo') or request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'Debes seleccionar un archivo CSV o XLSX'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = parse_career_bulk_file(uploaded_file)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    parsed_rows, rejected_rows = parse_career_catalog_rows(rows)
    return Response(
        {
            'rows': parsed_rows,
            'rejected_rows': rejected_rows,
            'loaded': len(parsed_rows),
            'rejected': len(rejected_rows),
            'skipped': len(rejected_rows),
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_career_bulk_upload(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    preview_rows = request.data.get('rows') if hasattr(request, 'data') else None
    if preview_rows is not None:
        created, updated = save_career_catalog_rows(preview_rows)
        return Response(
            {'message': 'Cargue finalizado', 'created': created, 'updated': updated, 'skipped': 0},
            status=status.HTTP_200_OK,
        )

    uploaded_file = request.FILES.get('archivo') or request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'Debes seleccionar un archivo CSV o XLSX'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = parse_career_bulk_file(uploaded_file)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    created = 0
    updated = 0
    skipped = 0

    with connection.cursor() as cursor:
        ensure_careers_active_column(cursor)
        cursor.execute('SELECT id, LOWER(nom_area) AS area FROM tb_area')
        area_by_name = {row['area']: row['id'] for row in dictfetchall(cursor)}

        for row in rows:
            normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
            nombre = str(normalized.get('nombre') or normalized.get('carrera') or '').strip()
            id_area = normalized.get('id_area') or normalized.get('area_id')
            area_name = str(normalized.get('area') or normalized.get('área') or '').strip().lower()
            activa = normalize_bool(normalized.get('estado', normalized.get('activa', True)))

            if not id_area and area_name:
                id_area = area_by_name.get(area_name)

            if not nombre or not id_area:
                skipped += 1
                continue

            cursor.execute('SELECT ID FROM carreras WHERE LOWER(Nom_carrera) = LOWER(%s)', [nombre])
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    'UPDATE carreras SET ID_area = %s, activa = %s WHERE ID = %s',
                    [id_area, 1 if activa else 0, existing[0]],
                )
                updated += 1
            else:
                cursor.execute(
                    'INSERT INTO carreras (Nom_carrera, ID_area, activa) VALUES (%s, %s, %s)',
                    [nombre, id_area, 1 if activa else 0],
                )
                created += 1

    return Response(
        {'message': 'Cargue finalizado', 'created': created, 'updated': updated, 'skipped': skipped},
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_entity_create(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    data = parse_json_request(request)
    nombre = (data.get('nombre') or '').strip()
    if not nombre:
        return Response({'error': 'El nombre es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)

    values = [
        nombre,
        (data.get('tipo') or '').strip() or None,
        (data.get('descripcion') or '').strip() or None,
        (data.get('contacto') or '').strip() or None,
        (data.get('sitio_web') or '').strip() or None,
    ]
    with connection.cursor() as cursor:
        cursor.execute(
            '''
            INSERT INTO entidades_apoyo (nombre, tipo, descripcion, contacto, sitio_web)
            VALUES (%s, %s, %s, %s, %s)
            ''',
            values,
        )
        entity_id = cursor.lastrowid

    return Response({'id': entity_id, **data}, status=status.HTTP_201_CREATED)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_admin_entity_detail(request, entity_id):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    with connection.cursor() as cursor:
        if request.method == 'DELETE':
            cursor.execute('DELETE FROM convenios WHERE id_entidad = %s', [entity_id])
            cursor.execute('DELETE FROM entidades_apoyo WHERE id = %s', [entity_id])
            return Response({'message': 'Entidad eliminada'}, status=status.HTTP_200_OK)

        data = parse_json_request(request)
        nombre = (data.get('nombre') or '').strip()
        if not nombre:
            return Response({'error': 'El nombre es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)

        cursor.execute(
            '''
            UPDATE entidades_apoyo
            SET nombre = %s, tipo = %s, descripcion = %s, contacto = %s, sitio_web = %s
            WHERE id = %s
            ''',
            [
                nombre,
                (data.get('tipo') or '').strip() or None,
                (data.get('descripcion') or '').strip() or None,
                (data.get('contacto') or '').strip() or None,
                (data.get('sitio_web') or '').strip() or None,
                entity_id,
            ],
        )

    return Response({'id': entity_id, **data}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_admin_university_create(request):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    data = parse_json_request(request)
    nombre = (data.get('nombre') or '').strip()
    tipo = (data.get('tipo') or '').strip()
    if not nombre or not tipo:
        return Response({'error': 'Nombre y tipo de universidad son obligatorios'}, status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        with connection.cursor() as cursor:
            cursor.execute(
                '''
                INSERT INTO universidades (nombre, tipo, localidad, direccion, sitio_web)
                VALUES (%s, %s, %s, %s, %s)
                ''',
                [
                    nombre,
                    tipo,
                    (data.get('localidad') or '').strip() or None,
                    (data.get('direccion') or '').strip() or None,
                    (data.get('sitio_web') or '').strip() or None,
                ],
            )
            university_id = cursor.lastrowid
            sync_university_relations(cursor, university_id, data)

    university = next((item for item in serialize_university_rows() if item['id'] == university_id), None)
    return Response(university or {'id': university_id}, status=status.HTTP_201_CREATED)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_admin_university_detail(request, university_id):
    _, error_response = ensure_admin_user(request)
    if error_response:
        return error_response

    with transaction.atomic():
        with connection.cursor() as cursor:
            if request.method == 'DELETE':
                cursor.execute('DELETE FROM beneficios_carrera WHERE id_universidad = %s', [university_id])
                cursor.execute('DELETE FROM convenios WHERE id_universidad = %s', [university_id])
                cursor.execute('DELETE FROM carrera_universidad WHERE id_universidad = %s', [university_id])
                cursor.execute('DELETE FROM universidades WHERE id = %s', [university_id])
                return Response({'message': 'Universidad eliminada'}, status=status.HTTP_200_OK)

            data = parse_json_request(request)
            nombre = (data.get('nombre') or '').strip()
            tipo = (data.get('tipo') or '').strip()
            if not nombre or not tipo:
                return Response({'error': 'Nombre y tipo de universidad son obligatorios'}, status=status.HTTP_400_BAD_REQUEST)

            cursor.execute(
                '''
                UPDATE universidades
                SET nombre = %s, tipo = %s, localidad = %s, direccion = %s, sitio_web = %s
                WHERE id = %s
                ''',
                [
                    nombre,
                    tipo,
                    (data.get('localidad') or '').strip() or None,
                    (data.get('direccion') or '').strip() or None,
                    (data.get('sitio_web') or '').strip() or None,
                    university_id,
                ],
            )
            sync_university_relations(cursor, university_id, data)

    university = next((item for item in serialize_university_rows() if item['id'] == university_id), None)
    return Response(university or {'id': university_id}, status=status.HTTP_200_OK)
